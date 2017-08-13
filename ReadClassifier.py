#!/usr/bin/env python
import json
from openpyxl import Workbook

data = []
data1 = []
mypath1 = '/home/aan/congestion/classification_name_17.json'
with open(mypath1) as file:
    data1 = json.load(file)
data2 = []
mypath2 = '/home/aan/congestion/classification_name_49.json'
with open(mypath2) as file:
    data2 = json.load(file)
data = data1 + data2

wb = Workbook()
dest_filename = 'data_train1.xlsx'
ws = wb.active
ws.title = 'Datum'

# Table Header
ws['A1'] = 'Nomor'
ws['B1'] = 'Responden'
ws['C1'] = 'Twit'
ws['D1'] = 'Klasifikasi'
ws['E1'] = 'Raw ID'

no = 1
row = 2
for d in data:
    ws['A' + str(row)] = no
    ws['B' + str(row)] = d['respondent']
    ws['C' + str(row)] = d['info']
    ws['D' + str(row)] = d['class']
    ws['E' + str(row)] = d['raw_id']
    no += 1
    row += 1
    print(d)
wb.save(dest_filename)

# Read Tags
#mypath3 = '/home/aan/congestion/word_tag.json'
