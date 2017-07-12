#!/usr/bin/env python
import json
from openpyxl import Workbook

data = []
mypath = '/home/aan/congestion/word_tag.json'
with open(mypath) as file:
    data = json.load(file)
wb = Workbook()
dest_filename = 'data_tag1.xlsx'
ws = wb.active
ws.title = 'Datum'

# Table Header
ws['A1'] = 'Nomor'
ws['B1'] = 'Word'
ws['C1'] = 'Tag'

no = 1
row = 2
words = []
for d in data:
    for t in d['tags']:
        ws['A' + str(row)] = no
        ws['B' + str(row)] = t[0]
        ws['C' + str(row)] = t[1]
        no += 1
        row += 1
        print(d)
        if t[0] not in words:
            words.append(t[0])
#wb.save(dest_filename)
print(len(words))
# Read Tags
#mypath3 = '/home/aan/congestion/word_tag.json'
