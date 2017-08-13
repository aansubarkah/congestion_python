#!/usr/bin/env python
from prettytable import PrettyTable
from debe import *
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_wb25.xlsx'
ws = wb.active
ws.title = 'Datum'

# Read Raws
query = sessionPostgresTraffic.query(DataTwitter).\
    filter(DataTwitter.at_classification_id != None).\
    limit(200)

# Table Header
ws['A1'] = 'Nomor'
ws['B1'] = 'Responden'
ws['C1'] = 'Twit'
ws['D1'] = 'Waktu'
ws['E1'] = 'Klasifikasi'
ws['E2'] = 'AT'
ws['F2'] = 'MT'
ws['G2'] = 'Benar/Salah'
ws['H1'] = 'Label Kata'
ws['I1'] = 'Lokasi'
ws['J1'] = 'Kondisi'
ws['K1'] = 'Cuaca'
ws['L1'] = 'Lokasi'
ws['M1'] = 'Kondisi'
ws['N1'] = 'Cuaca'
ws['O1'] = 'Raw ID'

i = 1
row = 3
for q in query:
    classificationTF = 0
    if q.at_classification_id == q.mt_classification_id:
        classificationTF = 1

    # Tagging
    queryTagging = sessionPostgresTraffic.query(DataWord).\
        filter(DataWord.raw_id == q.raw_id).\
        all()
    words = ''
    for d in queryTagging:
        words = words + ' ' + d.word_name + '/' + d.tag_name
    # Chunking
    queryChunking = sessionPostgresTraffic.query(DataChunk).\
        filter(DataChunk.raw_id == q.raw_id).\
        all()

    ws['A' + str(row)] = i
    ws['B' + str(row)] = q.respondent_name
    ws['C' + str(row)] = q.info
    ws['D' + str(row)] = q.t_time
    ws['E' + str(row)] = q.at_classification_id
    ws['F' + str(row)] = q.mt_classification_id
    ws['G' + str(row)] = classificationTF
    ws['H' + str(row)] = words
    ws['O' + str(row)] = q.raw_id

    for d in queryChunking:
        # Spoting
        querySpoting = sessionPostgresTraffic.query(DataSpot).\
            filter(DataSpot.chunk_id == d.chunk_id).\
            all()

        ws['I' + str(row)] = d.place
        ws['J' + str(row)] = d.condition
        ws['K' + str(row)] = d.weather

        for s in querySpoting:
            ws['L' + str(row)] = s.place_name
            ws['M' + str(row)] = s.category_name
            ws['N' + str(row)] = s.weather_name

        row += 1
    i += 1
    row += 1

wb.save(dest_filename)
